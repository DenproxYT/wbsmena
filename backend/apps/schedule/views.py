from rest_framework import generics, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date as date_type
from calendar import monthrange
from rest_framework.exceptions import PermissionDenied
from decimal import Decimal, InvalidOperation
from django.db.models import Max
import re
from .models import Schedule, HouseholdSupplyRequest
from .serializers import ScheduleSerializer, HouseholdSupplyRequestSerializer
from .constants import PVZ_ADDRESSES, UNIVERSAL_PVZ_LABEL
from .period_utils import (
    check_schedule_write_allowed,
    get_visible_months_for_user,
    ensure_period,
)
from .signals_schedule import emit_schedule_changed

User = get_user_model()


def is_schedule_admin(user):
    """Администратор или владелец: могут выбирать всех сотрудников и все/отдельные ПВЗ."""
    if not user or not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    return getattr(user, 'role', None) in ('administrator', 'owner')


def resolve_schedule_pvz_for_user(target_user, requested_pvz=''):
    """ПВЗ для записи графика: у универсалов — единая метка, иначе из запроса или профиля."""
    if getattr(target_user, 'is_universal', False):
        return UNIVERSAL_PVZ_LABEL
    pvz = (requested_pvz or '').strip()
    if pvz and pvz in PVZ_ADDRESSES:
        return pvz
    return (getattr(target_user, 'pvz_address', '') or '').strip() or (PVZ_ADDRESSES[0] if PVZ_ADDRESSES else '')


def _normalize_person_text(value: str) -> str:
    value = (value or "").lower().strip()
    value = value.replace("ё", "е")
    value = re.sub(r"\+?\d[\d\-\(\)\s]{8,}", " ", value)
    value = re.sub(r"[^a-zа-я0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _resolve_user_from_text(raw_value: str):
    src = (raw_value or "").strip()
    if not src:
        return None

    # Часть до "/" обычно содержит ФИО или логин.
    base = src.split("/")[0].strip()
    normalized = _normalize_person_text(base)
    if not normalized:
        return None

    # 1) Точный username.
    try:
        return User.objects.get(username=base)
    except User.DoesNotExist:
        pass

    # 2) Username без регистра/мусора.
    u = User.objects.filter(username__iexact=normalized).first()
    if u:
        return u

    # 3) ФИО по токенам.
    tokens = [t for t in normalized.split(" ") if t]
    if len(tokens) >= 2:
        cand = User.objects.filter(last_name__icontains=tokens[0], first_name__icontains=tokens[1]).first()
        if cand:
            return cand

    # 4) Более мягкий поиск по ФИО/username.
    q = User.objects.all()
    best = None
    best_score = 0
    for usr in q:
        username_n = _normalize_person_text(usr.username or "")
        fio_n = _normalize_person_text(f"{usr.last_name or ''} {usr.first_name or ''}")
        score = 0
        for t in tokens:
            if len(t) < 2:
                continue
            if t in username_n:
                score += 1
            if t in fio_n:
                score += 2
        if fio_n and normalized in fio_n:
            score += 3
        if username_n and normalized in username_n:
            score += 2
        if score > best_score:
            best_score = score
            best = usr
    return best if best_score >= 2 else None


def assert_can_write_schedule(request, schedule_date):
    if isinstance(schedule_date, str):
        schedule_date = date_type.fromisoformat(schedule_date[:10])
    ok, msg = check_schedule_write_allowed(
        request.user, schedule_date, is_schedule_admin(request.user)
    )
    if not ok:
        raise PermissionDenied(msg)


class ScheduleListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ScheduleSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Schedule.objects.select_related('user')
        
        # Фильтрация по дате
        date_year = self.request.query_params.get('date__year', None)
        date_month = self.request.query_params.get('date__month', None)
        if date_year:
            queryset = queryset.filter(date__year=date_year)
        if date_month:
            queryset = queryset.filter(date__month=date_month)
        
        # Фильтрация по пользователю (для админа/владельца)
        user_id = self.request.query_params.get('user', None)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        # Фильтрация по ПВЗ (для админа/владельца)
        pvz_address = self.request.query_params.get('pvz_address', None)
        if pvz_address and pvz_address.strip() and is_schedule_admin(user):
            queryset = queryset.filter(pvz_address=pvz_address.strip())
        # Ограничение доступа для обычных пользователей
        if not is_schedule_admin(user):
            queryset = queryset.filter(user=user)
        return queryset.order_by('-date', 'user__username')

    def _assert_can_write(self, schedule_date):
        assert_can_write_schedule(self.request, schedule_date)

    def perform_create(self, serializer):
        self._assert_can_write(serializer.validated_data['date'])
        user = self.request.user
        pvz = self.request.data.get('pvz_address', '').strip()
        
        if not is_schedule_admin(user):
            pvz_address = resolve_schedule_pvz_for_user(user, pvz)
            schedule = serializer.save(user=user, pvz_address=pvz_address)
        else:
            user_id = self.request.data.get('user')
            if user_id:
                try:
                    target_user = User.objects.get(id=user_id)
                    pvz_address = resolve_schedule_pvz_for_user(target_user, pvz)
                    schedule = serializer.save(user=target_user, pvz_address=pvz_address)
                except User.DoesNotExist:
                    pvz_address = resolve_schedule_pvz_for_user(user, pvz)
                    schedule = serializer.save(user=user, pvz_address=pvz_address)
            else:
                pvz_address = resolve_schedule_pvz_for_user(user, pvz)
                schedule = serializer.save(user=user, pvz_address=pvz_address)

        emit_schedule_changed(schedule.date)

        from apps.core.notification_service import broadcast_to_users
        from .period_utils import ensure_period as _ensure_period

        period = _ensure_period(schedule.date.year, schedule.date.month)
        period.updated_by = self.request.user
        period.save()

        if schedule.user and schedule.user.is_active and schedule.user != self.request.user:
            label = schedule.date.strftime('%d.%m.%Y')
            broadcast_to_users(
                title=f'Смена обновлена: {label}',
                message=(
                    f'На {label} изменены смены: {schedule.shifts}. '
                    f'ПВЗ: {(schedule.pvz_address or "").strip() or "—"}.'
                ),
                users=[schedule.user],
                created_by=self.request.user,
            )


class ScheduleRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ScheduleSerializer

    def get_queryset(self):
        user = self.request.user
        if is_schedule_admin(user):
            return Schedule.objects.all()
        return Schedule.objects.filter(user=user)

    def perform_destroy(self, instance):
        assert_can_write_schedule(self.request, instance.date)
        from apps.core.notification_service import broadcast_to_users
        from .period_utils import ensure_period as _ensure_period

        schedule_user = instance.user
        schedule_date = instance.date
        schedule_shifts = instance.shifts
        schedule_pvz_address = instance.pvz_address

        period = _ensure_period(schedule_date.year, schedule_date.month)
        period.updated_by = self.request.user
        period.save()

        instance.delete()
        emit_schedule_changed(schedule_date)

        if schedule_user and schedule_user.is_active and schedule_user != self.request.user:
            label = schedule_date.strftime('%d.%m.%Y')
            broadcast_to_users(
                title=f'Смена удалена: {label}',
                message=(
                    f'На {label} удалена смена: {schedule_shifts}. '
                    f'ПВЗ: {(schedule_pvz_address or "").strip() or "—"}.'
                ),
                users=[schedule_user],
                created_by=self.request.user,
            )

    def perform_update(self, serializer):
        d = serializer.validated_data.get('date', serializer.instance.date)
        assert_can_write_schedule(self.request, d)
        user = self.request.user
        pvz = self.request.data.get('pvz_address', '').strip()
        target = serializer.instance.user
        
        if not is_schedule_admin(user):
            pvz_address = resolve_schedule_pvz_for_user(user, pvz)
            schedule = serializer.save(user=user, pvz_address=pvz_address)
        else:
            user_id = self.request.data.get('user')
            if user_id:
                target_user = User.objects.get(id=user_id)
                target = target_user
            pvz_address = resolve_schedule_pvz_for_user(target, pvz)
            if user_id:
                schedule = serializer.save(user=target_user, pvz_address=pvz_address)
            else:
                schedule = serializer.save(pvz_address=pvz_address)

        emit_schedule_changed(schedule.date)

        from apps.core.notification_service import broadcast_to_users
        from .period_utils import ensure_period as _ensure_period

        period = _ensure_period(schedule.date.year, schedule.date.month)
        period.updated_by = self.request.user
        period.save()

        if schedule.user and schedule.user.is_active and schedule.user != self.request.user:
            label = schedule.date.strftime('%d.%m.%Y')
            broadcast_to_users(
                title=f'Смена обновлена: {label}',
                message=(
                    f'На {label} изменены смены: {schedule.shifts}. '
                    f'ПВЗ: {(schedule.pvz_address or "").strip() or "—"}.'
                ),
                users=[schedule.user],
                created_by=self.request.user,
            )


class ScheduleExportView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ScheduleSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Schedule.objects.select_related('user')
        
        # Фильтрация по дате
        date_year = self.request.query_params.get('date__year', None)
        date_month = self.request.query_params.get('date__month', None)
        if date_year:
            queryset = queryset.filter(date__year=date_year)
        if date_month:
            queryset = queryset.filter(date__month=date_month)
        
        # Фильтрация по пользователю
        user_id = self.request.query_params.get('user', None)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        # Фильтрация по ПВЗ (для админа/владельца)
        pvz_address = self.request.query_params.get('pvz_address', None)
        if pvz_address and pvz_address.strip() and is_schedule_admin(user):
            queryset = queryset.filter(pvz_address=pvz_address.strip())
        if not is_schedule_admin(user):
            queryset = queryset.filter(user=user)
        return queryset.select_related("user").order_by('date', 'user__username')

    def get(self, request, *args, **kwargs):
        if not is_schedule_admin(request.user):
            return Response({"detail": "Нет доступа"}, status=status.HTTP_403_FORBIDDEN)
        queryset = self.get_queryset()

        # Берем месяц/год из фильтра. Если фильтр не передан — текущий месяц.
        now = datetime.now()
        try:
            year = int(request.query_params.get("date__year") or now.year)
        except (TypeError, ValueError):
            year = now.year
        try:
            month = int(request.query_params.get("date__month") or now.month)
        except (TypeError, ValueError):
            month = now.month
        if month < 1 or month > 12:
            month = now.month

        month_names_ru = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
        ]
        days_in_month = monthrange(year, month)[1]

        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_names_ru[month - 1]} {year}"

        # Стили под табельный вид.
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        pvz_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")
        alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        thin = Side(style="thin", color="C9C9C9")
        thick = Side(style="medium", color="9E9E9E")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=1, column=1, value="WB Смена").font = title_font
        ws.cell(
            row=3,
            column=1,
            value="ПВЗ / ФИО сотрудника / телефон",
        ).font = header_font
        ws.cell(
            row=3,
            column=2,
            value=f"Расчетный месяц: {month_names_ru[month - 1]} / {year} г.",
        ).font = header_font

        # Строка с днями месяца.
        # Формат как в вашем шаблоне:
        # дни 1–15, затем 4 колонки итогов, затем дни 16–(конец), затем 5 колонок итогов.
        def col_for_day(d: int) -> int:
            # A=1 — ФИО/телефон; day1=B=2
            # Вставка 4 колонок после дня 15 -> с 16-го дня смещение +4.
            return (d + 1) if d <= 15 else (d + 5)

        summary1_start = col_for_day(15) + 1
        summary2_start = col_for_day(days_in_month) + 1  # AK

        for day in range(1, days_in_month + 1):
            col = col_for_day(day)
            c = ws.cell(row=4, column=col, value=day)
            c.font = header_font
            c.alignment = center
            c.fill = header_fill

        # Заголовки итогов (по вашей логике)
        # 1-я половина: удержания вместо процентов
        summary_headers = ["Общее 1-15", "Удержания", "Выплата", "К выдаче"]
        for i, h in enumerate(summary_headers):
            cell = ws.cell(row=4, column=summary1_start + i, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill

        # 2-я половина: добавляем удержания отдельной колонкой (5 колонок)
        summary2_headers = [f"Общее 16-{days_in_month}", "Проценты", "Выплата", "Удержания", "К выдаче"]
        for i, h in enumerate(summary2_headers):
            cell = ws.cell(row=4, column=summary2_start + i, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill

        def employee_home_pvz(user):
            if getattr(user, "is_universal", False):
                return None
            return (getattr(user, "pvz_address", "") or "").strip() or "Без ПВЗ"

        def short_pvz_label(pvz: str) -> str:
            return re.sub(r"^ЧИТА[_\s]*", "", (pvz or "").strip(), flags=re.I).strip() or pvz

        # Штатные: смены на «своём» ПВЗ и отдельно по каждой подмене.
        grouped = {}
        grouped_universal = {}
        staff_by_user = {}
        for s in queryset:
            user_key = s.user_id
            u = s.user
            if getattr(u, "is_universal", False):
                grouped_universal.setdefault(user_key, {"user": u, "days": {}})
                grouped_universal[user_key]["days"][s.date.day] = float(s.shifts or 0)
                continue
            home_pvz = employee_home_pvz(u)
            sched_pvz = (s.pvz_address or "").strip() or home_pvz
            row = staff_by_user.setdefault(
                user_key,
                {"user": u, "home_days": {}, "subs": {}},
            )
            day = s.date.day
            shifts = float(s.shifts or 0)
            if sched_pvz == home_pvz:
                row["home_days"][day] = row["home_days"].get(day, 0) + shifts
            else:
                sub = row["subs"].setdefault(sched_pvz, {})
                sub[day] = sub.get(day, 0) + shifts

        for user_key, user_row in staff_by_user.items():
            home_pvz = employee_home_pvz(user_row["user"])
            grouped.setdefault(home_pvz, {})[user_key] = user_row

        row_idx = 6
        last_col = summary2_start + 4
        total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        total_font = Font(bold=True)

        total1_col = summary1_start + 0
        ded1_col = summary1_start + 1
        pay1_col = summary1_start + 2
        out1_col = summary1_start + 3
        total2_col = summary2_start + 0
        pct2_col = summary2_start + 1
        pay2_col = summary2_start + 2
        ded2_col = summary2_start + 3
        out2_col = summary2_start + 4

        def write_days_row(ri, label, days_dict, base_fill=None):
            """Строка смен с формулами выплаты; возвращает номер строки."""
            ws.cell(row=ri, column=1, value=label).alignment = left
            for day in range(1, days_in_month + 1):
                val = days_dict.get(day)
                if val and val > 0:
                    cell = ws.cell(row=ri, column=col_for_day(day), value=val)
                    cell.alignment = center
            for c in range(1, last_col + 1):
                cell = ws.cell(row=ri, column=c)
                if base_fill and not cell.fill.patternType:
                    cell.fill = base_fill
                cell.border = thin_border
            total_1_15 = sum((days_dict.get(d, 0) or 0) for d in range(1, min(15, days_in_month) + 1))
            total_16_end = sum((days_dict.get(d, 0) or 0) for d in range(16, days_in_month + 1))
            ws.cell(row=ri, column=total1_col, value=(total_1_15 or "")).alignment = center
            ws.cell(row=ri, column=ded1_col, value="").alignment = center
            ws.cell(row=ri, column=pay1_col, value=1900).alignment = center
            ws.cell(
                row=ri, column=out1_col,
                value=f"={get_column_letter(pay1_col)}{ri}*{get_column_letter(total1_col)}{ri}-{get_column_letter(ded1_col)}{ri}",
            ).alignment = center
            ws.cell(row=ri, column=total2_col, value=(total_16_end or "")).alignment = center
            ws.cell(row=ri, column=pct2_col, value="").alignment = center
            ws.cell(row=ri, column=pay2_col, value=1900).alignment = center
            ws.cell(row=ri, column=ded2_col, value="").alignment = center
            ws.cell(
                row=ri, column=out2_col,
                value=(
                    f"=({get_column_letter(pay2_col)}{ri}*{get_column_letter(total2_col)}{ri})"
                    f"+({get_column_letter(pct2_col)}{ri}*({get_column_letter(total1_col)}{ri}+{get_column_letter(total2_col)}{ri}))"
                    f"-{get_column_letter(ded2_col)}{ri}"
                ),
            ).alignment = center
            return ri

        def write_employee_total_row(ri, payout_row_nums):
            ws.cell(row=ri, column=1, value="Итого").font = total_font
            ws.cell(row=ri, column=1).alignment = left
            for c in range(1, last_col + 1):
                cell = ws.cell(row=ri, column=c)
                cell.border = thin_border
            refs1 = ",".join(f"{get_column_letter(out1_col)}{r}" for r in payout_row_nums)
            refs2 = ",".join(f"{get_column_letter(out2_col)}{r}" for r in payout_row_nums)
            c1 = ws.cell(row=ri, column=out1_col, value=f"=SUM({refs1})")
            c1.alignment = center
            c1.font = total_font
            c1.fill = total_fill
            c2 = ws.cell(row=ri, column=out2_col, value=f"=SUM({refs2})")
            c2.alignment = center
            c2.font = total_font
            c2.fill = total_fill
            return ri + 1

        def write_user_row(ri, user_row, i_user):
            u = user_row["user"]
            full_name = f"{u.last_name or ''} {u.first_name or ''}".strip() or u.username
            phone = (getattr(u, "phone_number", "") or "").strip()
            base_fill = alt_fill if (i_user % 2 == 1) else None

            if getattr(u, "is_universal", False):
                label = f"{full_name} / {phone} [Универсал]"
                ri = write_days_row(ri, label, user_row.get("days") or {}, base_fill)
                return ri + 1

            base_label = f"{full_name} / {phone}"
            payout_rows = []
            has_home = any((user_row.get("home_days") or {}).values())
            if has_home or not user_row.get("subs"):
                ri = write_days_row(ri, base_label, user_row.get("home_days") or {}, base_fill)
                payout_rows.append(ri)
                ri += 1

            for sub_pvz in sorted((user_row.get("subs") or {}).keys()):
                sub_days = user_row["subs"][sub_pvz]
                if not any(sub_days.values()):
                    continue
                sub_label = f"{base_label} [подмена: {short_pvz_label(sub_pvz)}]"
                ri = write_days_row(ri, sub_label, sub_days, base_fill)
                payout_rows.append(ri)
                ri += 1

            if len(payout_rows) > 1:
                ri = write_employee_total_row(ri, payout_rows)
            return ri

        for pvz, users_map in grouped.items():
            # Строка ПВЗ — заливка на всю ширину + толстая граница сверху
            ws.cell(row=row_idx, column=1, value=f"ЧИТА_ {pvz}").font = header_font
            for c in range(1, last_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = pvz_fill
                cell.alignment = left if c == 1 else center
                cell.border = Border(left=thin, right=thin, top=thick, bottom=thin)
            row_idx += 1

            # Стабильный порядок сотрудников.
            users_rows = sorted(
                users_map.values(),
                key=lambda v: (
                    (v["user"].last_name or "").lower(),
                    (v["user"].first_name or "").lower(),
                    (v["user"].username or "").lower(),
                ),
            )
            for i_user, user_row in enumerate(users_rows):
                row_idx = write_user_row(row_idx, user_row, i_user)

            for c in range(1, last_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
            row_idx += 2

        if grouped_universal:
            ws.cell(row=row_idx, column=1, value=f"УНИВЕРСАЛЫ — {UNIVERSAL_PVZ_LABEL}").font = header_font
            for c in range(1, last_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.alignment = left if c == 1 else center
                cell.border = Border(left=thin, right=thin, top=thick, bottom=thin)
            row_idx += 1
            uni_rows = sorted(
                grouped_universal.values(),
                key=lambda v: (
                    (v["user"].last_name or "").lower(),
                    (v["user"].first_name or "").lower(),
                    (v["user"].username or "").lower(),
                ),
            )
            for i_user, user_row in enumerate(uni_rows):
                row_idx = write_user_row(row_idx, user_row, i_user)
            for c in range(1, last_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
            row_idx += 2

        # Ширины.
        # Ширины колонок: дни делаем узкими, чтобы табель не был “на полэкрана”.
        ws.column_dimensions["A"].width = 46
        for d in range(1, days_in_month + 1):
            ws.column_dimensions[get_column_letter(col_for_day(d))].width = 4.2

        # Итоги 1-15: Общее / Удержания / Выплата / К выдаче
        ws.column_dimensions[get_column_letter(summary1_start + 0)].width = 11.7109375
        ws.column_dimensions[get_column_letter(summary1_start + 1)].width = 11.5703125
        ws.column_dimensions[get_column_letter(summary1_start + 2)].width = 11.7109375
        ws.column_dimensions[get_column_letter(summary1_start + 3)].width = 13.28515625

        # Итоги 16-конец: Общее / Проценты / Выплата / Удержания / К выдаче
        ws.column_dimensions[get_column_letter(summary2_start + 0)].width = 12
        ws.column_dimensions[get_column_letter(summary2_start + 1)].width = 11.5703125
        ws.column_dimensions[get_column_letter(summary2_start + 2)].width = 7.28515625
        ws.column_dimensions[get_column_letter(summary2_start + 3)].width = 11.5703125
        ws.column_dimensions[get_column_letter(summary2_start + 4)].width = 9

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"tabel_wb_{year}_{month:02d}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class ScheduleImportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_schedule_admin(request.user):
            return Response({"detail": "Нет доступа"}, status=status.HTTP_403_FORBIDDEN)
        user = request.user
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Файл не был загружен'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Проверка расширения файла
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Поддерживаются только файлы Excel (.xlsx, .xls)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Загружаем Excel файл
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Определяем месяц и год из файла (ищем в первых строках)
            month_year = None
            month = None
            year = None
            
            for row_num in range(1, min(6, ws.max_row + 1)):
                for cell in ws[row_num]:
                    if cell.value:
                        cell_text = str(cell.value).lower()
                        # Ищем строку типа "Февраль / 2026" или "Расчетный месяц: Февраль / 2026"
                        if 'расчетный месяц' in cell_text or any(m in cell_text for m in [
                            'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                            'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь'
                        ]):
                            # Извлекаем месяц и год
                            months_ru = {
                                'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4,
                                'май': 5, 'июнь': 6, 'июль': 7, 'август': 8,
                                'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
                            }
                            for m_name, m_num in months_ru.items():
                                if m_name in cell_text:
                                    month = m_num
                                    break
                            year_match = re.search(r'20\d{2}', cell_text)
                            if year_match:
                                year = int(year_match.group())
                            if month and year:
                                month_year = (year, month)
                                break
                if month_year:
                    break
            
            # Ищем строку с заголовками (может быть не первая строка)
            header_row = 1
            headers = []
            date_columns = {}  # Словарь: номер колонки -> день месяца
            
            # Пробуем найти заголовки в первых 10 строках
            for row_num in range(1, min(11, ws.max_row + 1)):
                row_values = []
                for cell in ws[row_num]:
                    if cell.value:
                        row_values.append(str(cell.value).strip())
                    else:
                        row_values.append('')
                
                # Проверяем, есть ли номера дней (1-31) в заголовках
                day_numbers = []
                for idx, val in enumerate(row_values):
                    try:
                        day = int(val)
                        if 1 <= day <= 31:
                            day_numbers.append((idx, day))
                    except:
                        pass
                
                # Если нашли несколько номеров дней, это строка с датами
                if len(day_numbers) >= 5:  # Минимум 5 дней для уверенности
                    headers = row_values
                    header_row = row_num
                    for col_idx, day in day_numbers:
                        date_columns[col_idx] = day
                    break
            
            # Если не нашли формат с днями в заголовках, используем стандартный поиск
            if not date_columns:
                # Пробуем найти заголовки в первых 5 строках
                for row_num in range(1, min(6, ws.max_row + 1)):
                    row_values = []
                    for cell in ws[row_num]:
                        if cell.value:
                            row_values.append(str(cell.value).strip().lower())
                        else:
                            row_values.append('')
                    
                    # Проверяем, есть ли ключевые слова заголовков
                    row_text = ' '.join(row_values)
                    if any(keyword in row_text for keyword in ['дата', 'date', 'смен', 'shifts', 'фио', 'пользователь', 'user']):
                        headers = row_values
                        header_row = row_num
                        break
                
                if not headers:
                    # Если заголовки не найдены, используем первую строку
                    for cell in ws[1]:
                        headers.append(cell.value.lower() if cell.value else '')
                    header_row = 1
            
            # Определяем индексы колонок (более гибкий поиск)
            date_idx = None
            user_idx = None
            shifts_idx = None
            comment_idx = None
            pvz_idx = None
            
            for idx, header in enumerate(headers):
                if not header:
                    continue
                    
                header_lower = str(header).lower()
                
                # Поиск колонки с датой
                if date_idx is None and ('дата' in header_lower or 'date' in header_lower or 'день' in header_lower):
                    date_idx = idx
                
                # Поиск колонки с пользователем
                if user_idx is None and any(keyword in header_lower for keyword in [
                    'пользователь', 'user', 'username', 'фио', 'сотрудник', 
                    'работник', 'имя', 'name', 'логин', 'login'
                ]):
                    user_idx = idx
                
                # Поиск колонки с количеством смен
                if shifts_idx is None and any(keyword in header_lower for keyword in [
                    'смен', 'shifts', 'часов', 'hours', 'часы', 'количество смен',
                    'кол-во смен', 'смена', 'shift'
                ]):
                    shifts_idx = idx
                
                # Поиск колонки с комментарием
                if comment_idx is None and any(keyword in header_lower for keyword in [
                    'комментарий', 'comment', 'примечание', 'note', 'заметка'
                ]):
                    comment_idx = idx
                
                # Поиск колонки с ПВЗ
                if pvz_idx is None and any(keyword in header_lower for keyword in [
                    'пвз', 'pvz', 'адрес', 'address', 'точка', 'магазин'
                ]):
                    pvz_idx = idx
            
            # Если дата не найдена, пробуем найти по формату ячеек или позиции
            if date_idx is None:
                # Пробуем найти дату в первой колонке или колонках с датами
                for col_idx in range(min(5, len(headers))):
                    # Проверяем несколько строк после заголовка
                    for test_row in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
                        cell = ws.cell(row=test_row, column=col_idx + 1)
                        if cell.value:
                            if isinstance(cell.value, datetime):
                                date_idx = col_idx
                                break
                            elif isinstance(cell.value, str):
                                try:
                                    datetime.strptime(cell.value, '%d.%m.%Y')
                                    date_idx = col_idx
                                    break
                                except:
                                    try:
                                        datetime.strptime(cell.value, '%Y-%m-%d')
                                        date_idx = col_idx
                                        break
                                    except:
                                        pass
                    if date_idx is not None:
                        break
            
            # Если найден формат табеля с днями в заголовках (формат WB)
            if date_columns:
                # Обрабатываем данные в формате табеля WB
                created = 0
                updated = 0
                errors = []
                start_row = header_row + 1
                
                # Находим колонку с ФИО (обычно первая колонка A)
                name_col_idx = 0
                
                for row_num in range(start_row, ws.max_row + 1):
                    try:
                        # Получаем значение из колонки с именем
                        name_cell = ws.cell(row=row_num, column=name_col_idx + 1)
                        if not name_cell.value:
                            continue
                        
                        name_value = str(name_cell.value).strip()
                        
                        # Пропускаем служебные строки
                        if any(skip in name_value.lower() for skip in [
                            'передача смены',
                            'общее количество',
                            'итого',
                            'пвз №',
                            'фио',
                            'расчетный месяц',
                            'wb ',
                            'чита_',
                            'чита ',
                            'ул.',
                            'мкр.',
                            'с.',
                        ]):
                            continue
                        
                        # Извлекаем пользователя из ячейки табеля.
                        user_name = name_value.split('/')[0].strip()
                        if not user_name:
                            continue
                        
                        # Ищем пользователя
                        target_user = _resolve_user_from_text(name_value)
                        if target_user is None:
                            # Если пользователь не указан, используем текущего
                            if is_schedule_admin(user):
                                errors.append(f'Строка {row_num}: Пользователь "{user_name}" не найден')
                                continue
                            target_user = user
                        
                        # Проверка прав доступа
                        if not is_schedule_admin(user):
                            if target_user != user:
                                continue  # Пропускаем чужие записи
                        
                        # Извлекаем ПВЗ из имени (если есть)
                        pvz_value = target_user.pvz_address
                        if 'чит' in name_value.lower() or 'чита' in name_value.lower():
                            # Можно извлечь адрес ПВЗ из строки, но пока используем адрес пользователя
                            pass
                        
                        # Обрабатываем каждую колонку с датой
                        for col_idx, day in date_columns.items():
                            if col_idx >= len(list(ws[row_num])):
                                continue
                                
                            cell = ws.cell(row=row_num, column=col_idx + 1)
                            if not cell.value:
                                continue
                            
                            # Получаем значение смены
                            shifts_value = 0.0
                            try:
                                shifts_str = str(cell.value).replace(',', '.')  # Заменяем запятую на точку
                                shifts_value = float(shifts_str)
                                if shifts_value <= 0:
                                    continue  # Пропускаем нулевые значения
                            except (ValueError, TypeError):
                                # Пробуем извлечь число из строки
                                numbers = re.findall(r'\d+[,.]?\d*', str(cell.value))
                                if numbers:
                                    shifts_value = float(numbers[0].replace(',', '.'))
                                else:
                                    continue
                            
                            # Формируем дату
                            if month_year:
                                year, month_num = month_year
                                try:
                                    schedule_date = datetime(year, month_num, day).date()
                                except ValueError:
                                    # Если день не существует в месяце (например, 31 февраля)
                                    errors.append(f'Строка {row_num}, день {day}: Неверная дата для месяца {month_num}')
                                    continue
                            else:
                                # Если месяц не определен, используем текущий
                                today = datetime.now().date()
                                try:
                                    schedule_date = datetime(today.year, today.month, day).date()
                                except ValueError:
                                    continue
                            
                            # Создаем или обновляем запись
                            schedule, created_flag = Schedule.objects.update_or_create(
                                user=target_user,
                                date=schedule_date,
                                defaults={
                                    'shifts': Decimal(str(shifts_value)),
                                    'comment': '',
                                    'pvz_address': pvz_value
                                }
                            )
                            
                            if created_flag:
                                created += 1
                            else:
                                updated += 1
                    
                    except Exception as e:
                        errors.append(f'Строка {row_num}: Ошибка обработки - {str(e)}')
                        continue
                
                result = {
                    'created': created,
                    'updated': updated,
                    'errors': errors
                }
                
                if errors:
                    return Response(result, status=status.HTTP_207_MULTI_STATUS)
                
                return Response(result, status=status.HTTP_200_OK)
            
            # Стандартный формат (колонки с заголовками)
            if date_idx is None:
                return Response(
                    {'error': 'Не найдена колонка с датой. Убедитесь, что файл содержит колонку "Дата" или даты в первой колонке.'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Обрабатываем данные (начинаем со строки после заголовков)
            created = 0
            updated = 0
            errors = []
            start_row = header_row + 1
            
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
                try:
                    # Пропускаем пустые строки
                    if not row[date_idx].value:
                        continue
                    
                    # Дата
                    date_value = row[date_idx].value
                    schedule_date = None
                    
                    if isinstance(date_value, datetime):
                        schedule_date = date_value.date()
                    elif isinstance(date_value, str):
                        date_str = date_value.strip()
                        # Пробуем разные форматы даты
                        date_formats = [
                            '%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d',
                            '%d.%m.%y', '%d/%m/%y', '%d-%m-%Y', '%Y.%m.%d'
                        ]
                        for fmt in date_formats:
                            try:
                                schedule_date = datetime.strptime(date_str, fmt).date()
                                break
                            except:
                                continue
                    elif date_value is not None:
                        # Пробуем преобразовать число в дату (Excel даты)
                        try:
                            if isinstance(date_value, (int, float)):
                                # Excel хранит даты как числа (дни с 1900-01-01)
                                from datetime import timedelta
                                excel_epoch = datetime(1899, 12, 30)
                                schedule_date = (excel_epoch + timedelta(days=int(date_value))).date()
                        except:
                            pass
                    
                    if schedule_date is None:
                        errors.append(f'Строка {row_num}: Неверный формат даты: {date_value}')
                        continue
                    
                    # Пользователь
                    target_user = None
                    
                    if user_idx is not None and user_idx < len(row) and row[user_idx].value:
                        user_value = str(row[user_idx].value).strip()
                        if user_value:
                            try:
                                target_user = User.objects.get(id=int(user_value))
                            except (ValueError, User.DoesNotExist):
                                target_user = _resolve_user_from_text(user_value)
                                if target_user is None:
                                    errors.append(f'Строка {row_num}: Пользователь "{user_value}" не найден')
                                    continue
                    
                    # Если пользователь не указан, используем текущего
                    if target_user is None:
                        if is_schedule_admin(user):
                            errors.append(f'Строка {row_num}: Для админа необходимо указать пользователя')
                            continue
                        target_user = user
                    
                    # Проверка прав доступа
                    if not is_schedule_admin(user):
                        if target_user != user:
                            errors.append(f'Строка {row_num}: Вы можете создавать смены только для себя')
                            continue
                    
                    # Количество смен
                    shifts_value = 0.0
                    if shifts_idx is not None and shifts_idx < len(row) and row[shifts_idx].value is not None:
                        try:
                            shifts_value = float(row[shifts_idx].value)
                            if shifts_value < 0:
                                shifts_value = 0.0
                        except (ValueError, TypeError):
                            # Если не число, пробуем извлечь число из строки
                            try:
                                shifts_str = str(row[shifts_idx].value)
                                numbers = re.findall(r'\d+\.?\d*', shifts_str)
                                if numbers:
                                    shifts_value = float(numbers[0])
                                else:
                                    shifts_value = 0.0
                            except:
                                errors.append(f'Строка {row_num}: Неверное значение количества смен: {row[shifts_idx].value}')
                                continue
                    # Если колонка смен не найдена, но есть данные, считаем что смена = 1
                    elif shifts_idx is None:
                        # Проверяем, есть ли вообще данные в строке (не пустая строка)
                        has_data = any(cell.value for cell in row if cell.value)
                        if has_data:
                            shifts_value = 1.0  # По умолчанию 1 смена
                    
                    # Комментарий
                    comment_value = ''
                    if comment_idx is not None and row[comment_idx].value:
                        comment_value = str(row[comment_idx].value).strip()
                    
                    # ПВЗ
                    pvz_value = target_user.pvz_address
                    if pvz_idx is not None and row[pvz_idx].value:
                        pvz_value = str(row[pvz_idx].value).strip()
                        # Для обычных пользователей проверяем, что ПВЗ совпадает
                        if not is_schedule_admin(user):
                            if pvz_value != user.pvz_address:
                                pvz_value = user.pvz_address
                    
                    # Создаем или обновляем запись
                    schedule, created_flag = Schedule.objects.update_or_create(
                        user=target_user,
                        date=schedule_date,
                        defaults={
                            'shifts': Decimal(str(shifts_value)),
                            'comment': comment_value,
                            'pvz_address': pvz_value
                        }
                    )
                    
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
                        
                except Exception as e:
                    errors.append(f'Строка {row_num}: Ошибка обработки - {str(e)}')
                    continue
            
            result = {
                'created': created,
                'updated': updated,
                'errors': errors
            }
            
            if errors:
                return Response(result, status=status.HTTP_207_MULTI_STATUS)
            
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Ошибка при обработке файла: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ScheduleMonthsView(APIView):
    """Доступные месяцы и статус открытия для графика."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        admin = is_schedule_admin(request.user)
        months = get_visible_months_for_user(request.user, admin)
        return Response({
            'months': months,
            'is_admin': admin,
        })


class SchedulePeriodToggleView(APIView):
    """Открыть/закрыть проставление смен для месяца (только админ)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, year, month):
        if not is_schedule_admin(request.user):
            return Response({'detail': 'Нет доступа'}, status=status.HTTP_403_FORBIDDEN)
        try:
            year, month = int(year), int(month)
            if month < 1 or month > 12:
                raise ValueError
        except (TypeError, ValueError):
            return Response({'detail': 'Некорректный месяц'}, status=status.HTTP_400_BAD_REQUEST)

        is_open = request.data.get('is_open')
        if is_open is None:
            period = ensure_period(year, month)
            is_open = not period.is_open
        else:
            is_open = bool(is_open)

        period = ensure_period(year, month)
        period.is_open = is_open
        period.updated_by = request.user
        period.save(update_fields=['is_open', 'updated_by', 'updated_at'])

        if is_open:
            from apps.core.notification_service import broadcast_to_employees
            from .period_utils import _month_label
            label = _month_label(year, month)
            broadcast_to_employees(
                title=f'График открыт: {label}',
                message=f'Администратор открыл проставление смен за {label}. Заполните график в разделе «График».',
                created_by=request.user,
            )

        from .period_utils import _period_dict
        return Response(_period_dict(period))


class ScheduleRealtimeMetaView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            year = int(request.query_params.get('date__year') or 0)
            month = int(request.query_params.get('date__month') or 0)
        except (TypeError, ValueError):
            return Response({'detail': 'Некорректная дата'}, status=status.HTTP_400_BAD_REQUEST)

        if year < 2000 or month < 1 or month > 12:
            return Response({'detail': 'Некорректная дата'}, status=status.HTTP_400_BAD_REQUEST)

        admin = is_schedule_admin(request.user)

        schedule_qs = Schedule.objects.select_related('user').filter(
            date__year=year,
            date__month=month,
        )
        if not admin:
            schedule_qs = schedule_qs.filter(user=request.user)
        else:
            user_id = request.query_params.get('user')
            pvz_address = request.query_params.get('pvz_address')
            if user_id:
                schedule_qs = schedule_qs.filter(user_id=user_id)
            if pvz_address and pvz_address.strip():
                schedule_qs = schedule_qs.filter(pvz_address=pvz_address.strip())

        schedule_updated_at = schedule_qs.aggregate(m=Max('updated_at')).get('m')

        period = ensure_period(year, month)
        period_updated_at = period.updated_at

        schedule_ts = int(schedule_updated_at.timestamp() * 1000) if schedule_updated_at else 0
        period_ts = int(period_updated_at.timestamp() * 1000) if period_updated_at else 0
        latest_ts = max(schedule_ts, period_ts)

        return Response({
            'latest_version': latest_ts,
            'schedule_updated_at': schedule_updated_at.isoformat() if schedule_updated_at else None,
            'period_updated_at': period_updated_at.isoformat() if period_updated_at else None,
            'is_open': bool(period.is_open),
        })


class HouseholdSupplyRequestListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = HouseholdSupplyRequestSerializer

    def get_queryset(self):
        user = self.request.user
        if is_schedule_admin(user):
            return HouseholdSupplyRequest.objects.select_related('user').all()
        return HouseholdSupplyRequest.objects.filter(user=user)


class HouseholdSupplyExportView(APIView):
    """Выгрузка заявок на хоз.нужды в Excel с полями для отметки (галочки)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_schedule_admin(request.user):
            return Response({"detail": "Нет доступа"}, status=status.HTTP_403_FORBIDDEN)

        qs = HouseholdSupplyRequest.objects.select_related("user").order_by("-created_at")
        pvz_filter = request.query_params.get("pvz_address", "").strip()
        if pvz_filter:
            qs = qs.filter(pvz_address=pvz_filter)

        wb = Workbook()
        ws = wb.active
        ws.title = "Хоз.нужды"

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        check_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin = Side(style="thin", color="C9C9C9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "№", "Дата", "Сотрудник", "ПВЗ",
            "Позиция 1", "☐", "Позиция 2", "☐", "Позиция 3", "☐",
            "Позиция 4", "☐", "Позиция 5", "☐", "Примечание",
        ]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row = 2
        for num, req in enumerate(qs, start=1):
            u = req.user
            name = f"{u.last_name or ''} {u.first_name or ''}".strip() or u.username
            items = [req.item_1, req.item_2, req.item_3, req.item_4, req.item_5]
            ws.cell(row=row, column=1, value=num).alignment = center
            ws.cell(row=row, column=2, value=req.created_at.strftime("%d.%m.%Y %H:%M")).alignment = center
            ws.cell(row=row, column=3, value=name).alignment = left
            ws.cell(row=row, column=4, value=req.pvz_address).alignment = left
            col = 5
            for item in items:
                ws.cell(row=row, column=col, value=item or "").alignment = left
                check_cell = ws.cell(row=row, column=col + 1, value="☐")
                check_cell.alignment = center
                check_cell.fill = check_fill
                check_cell.border = border
                col += 2
            ws.cell(row=row, column=15, value="").alignment = left
            for c in range(1, 16):
                ws.cell(row=row, column=c).border = border
            row += 1

        widths = [5, 16, 22, 28, 24, 5, 24, 5, 24, 5, 24, 5, 24, 5, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 28
        for r in range(2, row):
            ws.row_dimensions[r].height = 22

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="hoz_nuzhdy_{stamp}.xlsx"'
        return response
